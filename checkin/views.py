from django.shortcuts import render, redirect
from django.urls import reverse

from .models import *
from django.http import JsonResponse

from openpyxl import load_workbook

from .helpers import *
# from django.contrib.auth.hashers import make_password


# login
def login_page(request):
    # go to login page
    if request.method == 'GET':
        # if already logged in
        if 'login' in request.session:
            return redirect(reverse("checkin:events"))
        # login page
        return render(request, 'checkin/login.html')
            
    # login api
    if request.method == 'POST':
        # get data
        username = request.POST.get('username')
        mdp = request.POST.get('mdp')
        # encrypt
        # mdp = make_password(mdp)
        
        # check if exists
        user = User.objects.all().filter(username=username, mdp=mdp)
        print(user)
        # true
        if user:
            request.session['login'] = True
            return JsonResponse({
                'login' : 'Success',
            })
        # error
        else:
            return JsonResponse({
                'login' : 'Failed',
            })
  
# logout
@login_required
def logout_page(request):
    del request.session['login']
    return redirect(reverse('checkin:login'))

# home redirect to events
@login_required
def home(request):
    return redirect(reverse('checkin:events'))

# events
@login_required
def events_page(request):
    events = Event.objects.all()
    return render(request, 'checkin/events.html', {
        'events': events,
    })


# add evenement
@login_required
def add_event(request):
    # get to the page
    if request.method == 'GET':
        return render(request, 'checkin/add.html')
    
    # add event
    elif request.method == 'POST':
        # get data
        nom = request.POST.get('nom')
        lien = request.POST.get('lien')
        cp = request.POST.get('cp')
        cs = request.POST.get('cs')
        excel = request.FILES.get('excel')
        
        # check data
        if ((not nom) or (not lien) or (not cp) or (not cs) or (not excel)):
            return JsonResponse({
                "status" : "ERROR"
            })
            
        # create event
        event = Event(nom=nom, lien=lien, couleur_principale=cp, couleur_secondaire=cs)
        event.save()
            
        # convert excel to db
        wb = load_workbook(excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            # get data fom excel
            name = row[0].value
            prenom = row[1].value
            
            # error in excel
            if ((not name) or (not prenom)):
                # delete event
                event.delete()
                
                return JsonResponse({
                "status" : "ERROR"
            })
                
            # inscrit l'eleve
            particip = Participant(nom=name, prenom=prenom, event=event)
            particip.save()
            
            
        # end
        path = reverse('checkin:event', kwargs={'lien':lien})
        print(path)
        return JsonResponse({
            "status" : "OK",
            "lien": path,
        })


# page d'event
@login_required
def event_page(request, lien):
    event = Event.objects.all().filter(lien=lien)
    if not event:
        return redirect(reverse('checkin:events'))
    
    event = event[0]
    event.lien = event.lien.upper()
    return render(request, 'checkin/event.html', {
        'event': event,
    })

# api participants des events
@login_required
def event_api(request, lien):
    # if not post
    if request.method != 'POST':
        return JsonResponse({
            'error' : 'Not found',
        })
        
    # get event
    event = Event.objects.all().filter(lien=lien)
    
    # not found
    if not event:
        return JsonResponse({
            'error' : 'Event does not exist',
        })
        
    # get participants
    event = event[0]
    participants_querie = event.participants.all()
    participants = []
    
    # to json
    for participant in participants_querie:
        new = {
            'id': 'p' + str(participant.id),
            'nom': participant.nom,
            'prenom': participant.prenom,
            'present': participant.present,
        }
        participants.append(new)
    
    # send 
    return JsonResponse({
        'participants': participants,
    })
    
# changer presence des participants 
@login_required
def presence(request, lien):
    # if not post
    if request.method != 'POST':
        return JsonResponse({
            'error' : 'Not found',
        })
        
    # get event
    event = Event.objects.all().filter(lien=lien)
    
    # not found
    if not event:
        return JsonResponse({
            'error' : 'Event does not exist',
        })
    
    # get participant
    event = event[0]
    id = request.POST.get('id')
    participant = event.participants.all().filter(id=id)
    # not found
    if not participant:
        return JsonResponse({
            'error' : 'Participant does not exist',
        })
    # change presence
    participant = participant[0]
    if participant.present:
        participant.present = False
    else:
        participant.present = True
    # save
    participant.save()
    
    # return
    return JsonResponse({
        'result' : 'success',
    })